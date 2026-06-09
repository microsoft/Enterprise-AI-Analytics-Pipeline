"""
Module 10: pax_csv_export — CSV/Excel Export & Schema Discovery
================================================================
Migrated from: PAX_Purview_Audit_Log_Processor_v1.11.1.ps1
Source lines: L13441–13540, L14763–14932, L14990–15106, L15121–15203

Level: 2 (depends on pax_data_transform)

Provides all file I/O for output:
- Streaming CSV writer (open / write rows / close)
- Agent filter for record inclusion/exclusion
- CSV import to list-of-dicts (DataTable equivalent)
- List-of-dicts to list-of-dicts conversion (DataTable equivalent)
- Excel export via openpyxl
- Replay inline export (re-processes raw CSV through explosion)
- Unified replay header builder (auto-detects schemas across activity types)
- M365 usage wide header (delegates to unified)
- Entra users schema validation
- Unique string builder

External dependencies: openpyxl (optional, for Excel export)
Design: Uses stdlib logging.getLogger(__name__).
        CSV writer is instance-based (not script-scoped).

PS-to-Python Function Mapping
──────────────────────────────────────────────────────────────────────────
│ #  │ PS Function                 │ PS Line │ Python Function                  │
│────│────────────────────────────│─────────│──────────────────────────────────│
│ 76 │ Open-CsvWriter             │ 13441   │ CsvWriter.__init__ / open()      │
│ 77 │ Close-CsvWriter            │ 13456   │ CsvWriter.close()                │
│ 78 │ Write-CsvRows              │ 13457   │ CsvWriter.write_rows()           │
│ 79 │ Test-AgentFilter           │ 13540   │ test_agent_filter()              │
│ 80 │ Import-CsvToDataTable      │ 14763   │ import_csv_to_data_table()       │
│ 81 │ ConvertTo-DataTable        │ 14842   │ convert_to_data_table()          │
│ 82 │ Export-DataTableToExcel     │ 14893   │ export_data_table_to_excel()     │
│ 83 │ ConvertTo-UniqueString     │ 14926   │ convert_to_unique_string()       │
│ 84 │ Invoke-ReplayInlineExport  │ 14990   │ invoke_replay_inline_export()    │
│ 85 │ Get-UnifiedReplayHeader    │ 15106   │ get_unified_replay_header()      │
│ 86 │ Add-Paths (inner)          │ 15121   │ _add_paths() (inner helper)      │
│ 87 │ Get-M365UsageWideHeader    │ 15179   │ get_m365_usage_wide_header()     │
│ 88 │ Test-EntraUsersSchema      │ 15203   │ test_entra_users_schema()        │
│ 89 │ Save-CsvAtomic             │ 11530   │ save_csv_atomic()                │
│ 90 │ Get-AppendFileRecordIds    │ 9222    │ get_append_file_record_ids()     │
│ 91 │ ConvertTo-RecordIdExclusion│ 7383    │ convert_to_record_id_exclusion() │
│    │ $M365UsageBaseHeader       │ 15093   │ M365_USAGE_BASE_HEADER           │
│    │ $EntraUsersHeader          │ 15190   │ ENTRA_USERS_HEADER               │
──────────────────────────────────────────────────────────────────────────

Test Results
────────────
Run: (pending)
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# PS $M365UsageBaseHeader (L15093-15101)
M365_USAGE_BASE_HEADER: list[str] = [
    'RecordId', 'CreationDate', 'RecordType', 'Operation', 'UserId',
    'AuditData', 'AssociatedAdminUnits', 'AssociatedAdminUnitsNames',
    'CreationTime', 'Id', 'OrganizationId', 'ResultStatus', 'UserKey',
    'UserType', 'Version', 'Workload', 'ClientIP', 'ObjectId',
    'AzureActiveDirectoryEventType',
    'ExtendedProperties', 'ExtendedProperties.ResultStatusDetail',
    'ExtendedProperties.Name', 'ExtendedProperties.Value',
    'ExtendedProperties.UserAgent', 'ExtendedProperties.RequestType',
    'ModifiedProperties',
    'Actor', 'Actor.ID', 'Actor.Type',
    'ActorContextId', 'ActorIpAddress', 'InterSystemsId', 'IntraSystemId',
    'SupportTicketId',
    'Target', 'Target.ID', 'Target.Type',
    'TargetContextId', 'ApplicationId',
    'DeviceProperties', 'DeviceProperties.OS', 'DeviceProperties.Name',
    'DeviceProperties.Value', 'DeviceProperties.BrowserType',
    'DeviceProperties.SessionId',
    'ErrorNumber',
    'ExtendedProperties.KeepMeSignedIn',
    'DeviceProperties.Id', 'DeviceProperties.DisplayName',
    'DeviceProperties.TrustType',
    'ExtendedProperties.UserAuthenticationMethod',
    'DeviceProperties.IsCompliant', 'DeviceProperties.IsCompliantAndManaged',
    # SharePoint / OneDrive
    'SiteUrl', 'SourceRelativeUrl', 'SourceFileName', 'SourceFileExtension',
    'ListId', 'ListItemUniqueId', 'WebId', 'ApplicationDisplayName',
    'EventSource', 'ItemType', 'SiteSensitivityLabelId', 'GeoLocation',
    'IsManagedDevice', 'DeviceDisplayName', 'ListBaseType',
    'ListServerTemplate', 'AuthenticationType', 'Site',
    'DoNotDistributeEvent', 'HighPriorityMediaProcessing',
    # App Access Context
    'AppAccessContext.ClientAppId', 'AppAccessContext.ClientAppName',
    'AppAccessContext.CorrelationId', 'AppAccessContext.AADSessionId',
    'AppAccessContext.UniqueTokenId', 'AppAccessContext.AuthTime',
    'AppAccessContext.TokenIssuedAtTime', 'AppAccessContext.UserObjectId',
    'AppAccessContext.DeviceId',
]

# PS $EntraUsersHeader (L15190-15203) — 47 columns
ENTRA_USERS_HEADER: list[str] = [
    # Core Identity Properties
    'userPrincipalName', 'displayName', 'id', 'mail', 'givenName', 'surname',
    # Job Properties
    'jobTitle', 'department', 'employeeType', 'employeeId', 'employeeHireDate',
    # Location Properties
    'officeLocation', 'city', 'state', 'country', 'postalCode', 'companyName',
    # Organizational Properties (flattened from employeeOrgData)
    'employeeOrgData_division', 'employeeOrgData_costCenter',
    # Status Properties
    'accountEnabled', 'userType', 'createdDateTime',
    # Usage Properties
    'usageLocation', 'preferredLanguage',
    # Sync Properties
    'onPremisesSyncEnabled', 'onPremisesImmutableId', 'externalUserState',
    # Proxy Addresses (flattened)
    'proxyAddresses_Primary', 'proxyAddresses_Count', 'proxyAddresses_All',
    # Manager Properties (flattened from $expand=manager)
    'manager_id', 'manager_displayName', 'manager_userPrincipalName',
    'manager_mail', 'manager_jobTitle',
    # License Properties
    'assignedLicenses', 'hasLicense',
    # Power BI template compatibility columns (alias mappings)
    'ManagerID', 'BusinessAreaLabel', 'CountryofEmployment',
    'CompanyCodeLabel', 'CostCentreLabel', 'UserName',
    # Power BI template compatibility columns (null placeholders for Viva Insights fields)
    'EffectiveDate', 'FunctionType', 'BusinessAreaCode', 'OrgLevel_3Label',
]


# ═══════════════════════════════════════════════════════════════════════════════
# CSV WRITER — PS Open-CsvWriter / Close-CsvWriter / Write-CsvRows (L13441-13539)
# ═══════════════════════════════════════════════════════════════════════════════

# Pre-compiled regex for fields needing quoting (PS $needsQuotePattern)
_NEEDS_QUOTE_RE = re.compile(r'[",\r\n]')


class CsvWriter:
    """Streaming CSV writer matching PS Open-CsvWriter / Write-CsvRows / Close-CsvWriter.

    PS uses a 1 MB buffered StreamWriter + manual CSV escaping with a 4 MB StringBuilder.
    Python equivalent: buffered file with manual CSV formatting for exact PS parity.
    """

    def __init__(self, path: str, columns: list[str]):
        """Open CSV file and write header row.

        PS Open-CsvWriter (L13441-13455):
        - Creates UTF-8 StreamWriter (no BOM) with 1 MB buffer
        - Escapes column names: doubles quotes, wraps if contains [",\\r\\n] or leading/trailing space
        - Writes header line
        """
        self._path = path
        self._columns = list(columns)
        self._column_index: dict[str, int] = {c: i for i, c in enumerate(columns)}
        self._file = open(path, 'w', encoding='utf-8', newline='', buffering=1048576)
        # Write header row — PS Export-Csv quotes ALL fields unconditionally
        escaped = []
        for col in columns:
            c = str(col).replace('"', '""')
            escaped.append('"' + c + '"')
        self._file.write(','.join(escaped) + '\n')

    def close(self):
        """Close the CSV writer. PS Close-CsvWriter (L13456)."""
        if self._file:
            try:
                self._file.flush()
                self._file.close()
            except Exception:
                pass
            self._file = None

    def write_rows(self, rows, columns: Optional[list[str]] = None):
        """Write data rows to CSV.

        PS Write-CsvRows (L13457-13539):
        - Pre-compiles regex once
        - Uses 4 MB StringBuilder buffer
        - Builds column index lookup for O(1) access
        - For hashtables: iterates keys, maps to column index
        - For PSObjects: iterates properties, maps to column index
        - Handles arrays/collections -> semicolon-joined
        - Escapes: doubles quotes, wraps if contains [",\\r\\n] or leading/trailing space
        - Flushes at 4 MB
        """
        if rows is None:
            return
        if self._file is None:
            raise RuntimeError("CSV writer not initialized")

        cols = columns if columns is not None else self._columns
        col_count = len(cols)
        # Rebuild column index if custom columns provided
        col_index = self._column_index if columns is None else {c: i for i, c in enumerate(cols)}

        buf = []
        buf_size = 0

        for row in rows:
            if row is None:
                continue

            field_values = ['""'] * col_count  # PS Export-Csv: empty fields are quoted ""

            # PS: iterate keys of hashtable (or properties of PSObject)
            # Python: iterate dict keys
            if isinstance(row, dict):
                for key, val in row.items():
                    idx = col_index.get(key)
                    if idx is None:
                        continue
                    if val is None:
                        continue
                    # Handle arrays/collections -> semicolon-joined
                    # PS: if ($val -is [IEnumerable] -and -not ($val -is [string]))
                    if isinstance(val, (list, tuple, set)):
                        try:
                            val = ';'.join(str(v) if v is not None else '' for v in val)
                        except Exception:
                            val = str(val)
                    s = str(val)
                    # PS Export-Csv quotes ALL fields unconditionally
                    s = '"' + s.replace('"', '""') + '"'
                    field_values[idx] = s

            line = ','.join(field_values) + '\n'
            buf.append(line)
            buf_size += len(line)

            # Flush at 4 MB (PS flushes StringBuilder at 4194304)
            if buf_size > 4194304:
                self._file.write(''.join(buf))
                buf.clear()
                buf_size = 0

        if buf:
            self._file.write(''.join(buf))


# ═══════════════════════════════════════════════════════════════════════════════
# CONVERT TO DATA TABLE — PS ConvertTo-DataTable (L14842-14892)
# ═══════════════════════════════════════════════════════════════════════════════

def convert_to_data_table(input_objects: list[dict[str, Any]]) -> list[dict[str, str]]:
    """Convert list of dicts to list of string-valued dicts (DataTable equivalent).

    PS ConvertTo-DataTable (L14842-14892):
    - Creates DataTable from first row's properties
    - All columns are [string] type
    - None -> DBNull (Python: None)
    - All values cast to string
    """
    if not input_objects:
        return []

    # Get columns from first row
    columns = list(input_objects[0].keys())
    result: list[dict[str, str]] = []

    for obj in input_objects:
        row: dict[str, str] = {}
        for col in columns:
            val = obj.get(col)
            row[col] = None if val is None else str(val)
        result.append(row)

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT — PS Export-DataTableToExcel (L14893-14925)
# ═══════════════════════════════════════════════════════════════════════════════

def export_data_table_to_excel(
    data: list[dict[str, Any]],
    path: str,
    worksheet_name: str,
) -> None:
    """Export list of dicts to Excel using openpyxl.

    PS Export-DataTableToExcel (L14893-14925):
    - Converts data to DataTable via ConvertTo-DataTable
    - Uses Send-SQLDataToExcel with -FreezeTopRow -BoldTopRow -AutoSize -NoNumberConversion '*'
    - Python: openpyxl with freeze panes, bold header, auto-width
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
    except ImportError:
        logger.error("openpyxl is required for Excel export. Install with: pip install openpyxl")
        raise

    # Convert to string-typed data table
    table = convert_to_data_table(data)
    if not table:
        return

    columns = list(table[0].keys())

    # Load existing workbook or create new
    if os.path.exists(path):
        wb = load_workbook(path)
        if worksheet_name in wb.sheetnames:
            del wb[worksheet_name]
        ws = wb.create_sheet(worksheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = worksheet_name

    # Write header (PS: -BoldTopRow)
    bold_font = Font(bold=True)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold_font

    # Write data rows (PS: -NoNumberConversion '*' means all values are strings)
    for row_idx, row_data in enumerate(table, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')

    # PS: -FreezeTopRow
    ws.freeze_panes = 'A2'

    # PS: -AutoSize — approximate column width from content
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in table[:100]:  # Sample first 100 rows for width
            val = row.get(col_name, '')
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# UNIFIED REPLAY HEADER — PS Get-UnifiedReplayHeader (L15106-15177)
# ═══════════════════════════════════════════════════════════════════════════════

def _test_scalar_value(v: Any) -> bool:
    """Local scalar test — matches PS Test-ScalarValue logic.

    Used by _add_paths. Avoids import of mod2 for this single use.
    PS: returns $true for string, int, float, bool, datetime
    """
    return isinstance(v, (str, int, float, bool, type(None)))


def _add_paths(node: Any, prefix: str, collector: list[str]) -> None:
    """Recursively detect column paths from JSON, skipping CopilotEventData.

    PS Add-Paths inner function (L15121-15154):
    - None/scalar -> add prefix
    - Array -> recurse into each element
    - Dict -> recurse into properties
    - Skip CopilotEventData (handled by explosion)
    - Special handling for ExtendedProperties/DeviceProperties Name/Value arrays
    """
    if node is None:
        return
    if _test_scalar_value(node):
        if prefix:
            collector.append(prefix)
        return
    if isinstance(node, (list, tuple)) and not isinstance(node, str):
        for item in node:
            _add_paths(item, prefix, collector)
        return
    if isinstance(node, dict):
        for pn, pv in node.items():
            path = f"{prefix}.{pn}" if prefix else pn
            # SKIP CopilotEventData — explosion handles these with flat column names
            if pn == 'CopilotEventData':
                continue
            # Special handling for Name/Value arrays (pivot into columns)
            if pn == 'ExtendedProperties' and isinstance(pv, list):
                for item in pv:
                    try:
                        if isinstance(item, dict) and item.get('Name'):
                            collector.append(f"ExtendedProperties.{item['Name']}")
                    except Exception:
                        pass
                continue
            if pn == 'DeviceProperties' and isinstance(pv, list):
                for item in pv:
                    try:
                        if isinstance(item, dict) and item.get('Name'):
                            collector.append(f"DeviceProperties.{item['Name']}")
                    except Exception:
                        pass
                continue
            _add_paths(pv, path, collector)


def get_unified_replay_header(
    raw_csv_path: str,
    purview_exploded_header: list[str],
    sample: int = 500,
) -> list[str]:
    """Build unified replay header by scanning input CSV for all column paths.

    PS Get-UnifiedReplayHeader (L15106-15177):
    - Base columns common to all activity types (52 cols)
    - Augmented columns: SharePoint/OneDrive + AppAccessContext (32 cols)
    - Scans first N rows of raw CSV, parses AuditData JSON
    - Detects column paths recursively (skipping CopilotEventData)
    - Merges: base + aug + detected + PurviewExplodedHeader
    - Deduplicates (order-preserving)
    - Writes UnifiedReplayHeader.txt to raw CSV dir

    Parameters:
        raw_csv_path: Path to raw CSV file to scan
        purview_exploded_header: The 153-column PurviewExplodedHeader from mod9
        sample: Number of rows to scan (default 500)
    Returns:
        Unified header as list of column names
    """
    # PS $base (L15112)
    base = [
        'RecordId', 'CreationDate', 'RecordType', 'Operation', 'UserId',
        'AuditData', 'AssociatedAdminUnits', 'AssociatedAdminUnitsNames',
        'CreationTime', 'Id', 'OrganizationId', 'ResultStatus', 'UserKey',
        'UserType', 'Version', 'Workload', 'ClientIP', 'ObjectId',
        'AzureActiveDirectoryEventType',
        'ExtendedProperties', 'ExtendedProperties.ResultStatusDetail',
        'ExtendedProperties.Name', 'ExtendedProperties.Value',
        'ExtendedProperties.UserAgent', 'ExtendedProperties.RequestType',
        'ModifiedProperties',
        'Actor', 'Actor.ID', 'Actor.Type',
        'ActorContextId', 'ActorIpAddress', 'InterSystemsId', 'IntraSystemId',
        'SupportTicketId',
        'Target', 'Target.ID', 'Target.Type',
        'TargetContextId', 'ApplicationId',
        'DeviceProperties', 'DeviceProperties.OS', 'DeviceProperties.Name',
        'DeviceProperties.Value', 'DeviceProperties.BrowserType',
        'DeviceProperties.SessionId',
        'ErrorNumber',
        'ExtendedProperties.KeepMeSignedIn',
        'DeviceProperties.Id', 'DeviceProperties.DisplayName',
        'DeviceProperties.TrustType',
        'ExtendedProperties.UserAuthenticationMethod',
        'DeviceProperties.IsCompliant', 'DeviceProperties.IsCompliantAndManaged',
    ]
    # PS $aug (L15116-15119)
    aug = [
        'SiteUrl', 'SourceRelativeUrl', 'SourceFileName', 'SourceFileExtension',
        'ListId', 'ListItemUniqueId', 'WebId', 'ApplicationDisplayName',
        'EventSource', 'ItemType', 'SiteSensitivityLabelId', 'GeoLocation',
        'IsManagedDevice', 'DeviceDisplayName', 'ListBaseType',
        'ListServerTemplate', 'AuthenticationType', 'Site',
        'DoNotDistributeEvent', 'HighPriorityMediaProcessing',
        'AppAccessContext.ClientAppId', 'AppAccessContext.ClientAppName',
        'AppAccessContext.CorrelationId', 'AppAccessContext.AADSessionId',
        'AppAccessContext.UniqueTokenId', 'AppAccessContext.AuthTime',
        'AppAccessContext.TokenIssuedAtTime', 'AppAccessContext.UserObjectId',
        'AppAccessContext.DeviceId',
        'AppAccessContext.@odata.type', 'AppAccessContext.APIId',
        'AppAccessContext.IssuedAtTime',
    ]

    detected: list[str] = []

    if raw_csv_path and os.path.isfile(raw_csv_path):
        try:
            with open(raw_csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                count = 0
                for r in reader:
                    if count >= sample:
                        break
                    try:
                        audit_str = r.get('AuditData', '')
                        if audit_str:
                            audit = json.loads(audit_str)
                            if audit:
                                _add_paths(audit, '', detected)
                    except Exception:
                        pass
                    count += 1
        except Exception:
            pass

    # Build unified header: base + aug + detected + PurviewExplodedHeader (dedup, order-preserving)
    header: list[str] = []
    seen: set[str] = set()
    for source in [base, aug, detected, purview_exploded_header]:
        for c in source:
            if c not in seen:
                seen.add(c)
                header.append(c)

    # Write header file (PS L15170-15175)
    try:
        if raw_csv_path:
            hdr_dir = os.path.dirname(raw_csv_path)
            if hdr_dir:
                hdr_path = os.path.join(hdr_dir, 'UnifiedReplayHeader.txt')
                with open(hdr_path, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(header))
    except Exception:
        pass

    return header


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRA USERS SCHEMA VALIDATION — PS Test-EntraUsersSchema (L15203-15218)
# ═══════════════════════════════════════════════════════════════════════════════

def test_entra_users_schema(
    users: list[dict[str, Any]],
    quiet: bool = False,
) -> dict:
    """Validate Entra user records against expected 47-column schema.

    PS Test-EntraUsersSchema (L15203-15218):
    - If empty -> return (no-op)
    - Compare first user's keys against $EntraUsersHeader
    - Log missing and extra columns as WARNING
    - If valid and not quiet, log validation success

    Returns:
        dict with 'missing' and 'extra' lists (empty if valid)
    """
    result = {'missing': [], 'extra': []}
    if not users or len(users) == 0:
        return result

    expected = ENTRA_USERS_HEADER
    actual = list(users[0].keys())

    missing = [c for c in expected if c not in actual]
    extra = [c for c in actual if c not in expected]
    result['missing'] = missing
    result['extra'] = extra

    if missing or extra:
        logger.warning(
            f"EntraUsers schema mismatch. Missing: {', '.join(missing)}; "
            f"Extra: {', '.join(extra)}"
        )
    elif not quiet:
        logger.info(f"Validated EntraUsers schema ({len(expected)} columns).")

    return result

